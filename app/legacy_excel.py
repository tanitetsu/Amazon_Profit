"""Load order rows from legacy local Excel (Amazon利益管理シート①)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.order_sku import normalize_sku
from app.schema import (
    STATUS_BUYER_CANCEL,
    STATUS_OPEN,
    STATUS_RETURN,
    points_fallback_from_price_tax,
)


@dataclass
class LegacyOrderRow:
    order_id: str
    sku: str
    title: str
    order_date: date
    ship_by: date | None
    price: float | None
    tax: float | None
    fee: float | None
    points: int | None
    proceeds: float | None
    cost: float | None
    ship_date: date | None
    shipped: bool
    status: str
    comment: str
    cancel_lock: bool  # lock キャンセル☑ (× or 返品)


def _as_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel serial (Windows 1900 system)
        try:
            return date(1899, 12, 30) + timedelta(days=int(v))
        except (OverflowError, ValueError):
            return None
    return None


def _as_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _as_bool01(v: Any) -> bool:
    if v in (1, 1.0, True, "1", "TRUE", "true"):
        return True
    return False


def load_legacy_orders(path: str | Path) -> list[LegacyOrderRow]:
    """
    Parse 注文データ sheet.
    Columns (1-based): 1注文番号 2SKU 3商品名 4注文日 5出荷予定 6販売価格 7税金
    8手数料 9売上金 10仕入れ 11利益 12利益率 13Pt 14発送通知日 16発送通知
    17Cancel 18返品 19コメント(任意)
    """
    wb = load_workbook(path, data_only=True)
    if "注文データ" not in wb.sheetnames:
        raise ValueError(f"sheet 注文データ not found in {path}")
    ws = wb["注文データ"]
    out: list[LegacyOrderRow] = []

    for r in range(4, (ws.max_row or 0) + 1):
        order_id = ws.cell(r, 1).value
        if not order_id:
            continue
        order_id = str(order_id).strip()
        if order_id == "注文番号":
            continue
        od = _as_date(ws.cell(r, 4).value)
        if od is None:
            continue

        price = _as_float(ws.cell(r, 6).value)
        tax = _as_float(ws.cell(r, 7).value)
        fee = _as_float(ws.cell(r, 8).value)
        proceeds = _as_float(ws.cell(r, 9).value)
        cost = _as_float(ws.cell(r, 10).value)
        if cost == 0:
            cost = None
        pt = _as_float(ws.cell(r, 13).value)
        points = (
            int(pt) if pt is not None else points_fallback_from_price_tax(price, tax)
        )

        cancelled = _as_bool01(ws.cell(r, 17).value)
        returned = _as_bool01(ws.cell(r, 18).value)
        if returned:
            status = STATUS_RETURN
            cancel_lock = True
        elif cancelled:
            status = STATUS_BUYER_CANCEL
            cancel_lock = True
        else:
            status = STATUS_OPEN
            cancel_lock = False

        comment_v = ws.cell(r, 19).value
        comment = str(comment_v).strip() if comment_v not in (None, "") else ""

        sku = normalize_sku(str(ws.cell(r, 2).value or ""))
        title = str(ws.cell(r, 3).value or "").strip()
        # Multi-item legacy junk companion (no title / money) after placeholder SKU row
        if not title and price is None and fee is None and proceeds is None:
            continue

        out.append(
            LegacyOrderRow(
                order_id=order_id,
                sku=sku,
                title=title,
                order_date=od,
                ship_by=_as_date(ws.cell(r, 5).value),
                price=price,
                tax=tax,
                fee=fee,
                points=points,
                proceeds=proceeds,
                cost=cost,
                ship_date=_as_date(ws.cell(r, 14).value),
                shipped=_as_bool01(ws.cell(r, 16).value),
                status=status,
                comment=comment,
                cancel_lock=cancel_lock,
            )
        )
    return out


def group_by_month(rows: list[LegacyOrderRow]) -> dict[str, list[LegacyOrderRow]]:
    grouped: dict[str, list[LegacyOrderRow]] = {}
    for row in rows:
        key = f"{row.order_date.year:04d}-{row.order_date.month:02d}"
        grouped.setdefault(key, []).append(row)
    for key in grouped:
        grouped[key].sort(key=lambda x: (x.order_date, x.order_id, x.sku))
    return grouped
